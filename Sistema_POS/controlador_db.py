import pandas as pd
import os
import ast
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA_DATA = os.path.join(BASE_DIR, "data")
ARCHIVO_DB = os.path.join(CARPETA_DATA, "inventario.xlsx")
ARCHIVO_CONFIG = os.path.join(CARPETA_DATA, "config.txt") # NUEVO ARCHIVO DE CONTRASEÑA

class ControladorDB:
    def __init__(self):
        if not os.path.exists(CARPETA_DATA):
            os.makedirs(CARPETA_DATA)
        self.verificar_archivo()
        self.reparar_encabezados()
        self.verificar_config() # Verifica la contraseña

    def verificar_config(self):
        # Si no existe el archivo de configuración, crea uno con la clave por defecto
        if not os.path.exists(ARCHIVO_CONFIG):
            with open(ARCHIVO_CONFIG, 'w') as f:
                f.write("ADMINCHUY")

    def obtener_password(self):
        try:
            with open(ARCHIVO_CONFIG, 'r') as f:
                return f.read().strip()
        except:
            return "ADMINCHUY"

    def cambiar_password(self, actual, nueva):
        if actual == self.obtener_password():
            with open(ARCHIVO_CONFIG, 'w') as f:
                f.write(nueva.strip())
            return True, "Contraseña actualizada exitosamente."
        return False, "La contraseña actual es incorrecta."

    def verificar_archivo(self):
        if not os.path.exists(ARCHIVO_DB):
            with pd.ExcelWriter(ARCHIVO_DB, engine='openpyxl') as writer:
                df_prod = pd.DataFrame(columns=["codigo", "nombre", "precio", "stock"])
                df_prod.to_excel(writer, sheet_name="Productos", index=False)
                df_ventas = pd.DataFrame(columns=["fecha", "total", "items"])
                df_ventas.to_excel(writer, sheet_name="Ventas", index=False)

    def reparar_encabezados(self):
        try:
            df = pd.read_excel(ARCHIVO_DB, sheet_name="Ventas")
            if 'fecha' not in df.columns:
                df_reparado = pd.read_excel(ARCHIVO_DB, sheet_name="Ventas", header=None)
                df_reparado.columns = ["fecha", "total", "items"]
                df_prod = pd.read_excel(ARCHIVO_DB, sheet_name="Productos")
                with pd.ExcelWriter(ARCHIVO_DB, engine='openpyxl') as writer:
                    df_prod.to_excel(writer, sheet_name="Productos", index=False)
                    df_reparado.to_excel(writer, sheet_name="Ventas", index=False)
        except:
            pass

    def buscar_producto(self, codigo):
        try:
            df = pd.read_excel(ARCHIVO_DB, sheet_name="Productos", dtype={'codigo': str})
            producto = df[df['codigo'].str.strip() == str(codigo).strip()]
            if not producto.empty:
                return producto.iloc[0].to_dict()
            return None
        except:
            return None

    def registrar_venta_db(self, total, items):
        try:
            wb = load_workbook(ARCHIVO_DB)
            ws_ventas = wb["Ventas"]
            ws_ventas.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), total, str(items)])
            
            if "Productos" in wb.sheetnames:
                ws_productos = wb["Productos"]
                for item in items:
                    cod = str(item['codigo']).strip()
                    for fila in range(2, ws_productos.max_row + 1):
                        if str(ws_productos.cell(row=fila, column=1).value).strip() == cod:
                            stock_actual = int(ws_productos.cell(row=fila, column=4).value or 0)
                            ws_productos.cell(row=fila, column=4, value=stock_actual - 1)
                            break
            wb.save(ARCHIVO_DB)
            return True
        except:
            return False

    def obtener_ventas_por_fecha(self, fecha_busqueda):
        try:
            df = pd.read_excel(ARCHIVO_DB, sheet_name="Ventas")
            df['fecha'] = df['fecha'].astype(str)
            df['fecha_simple'] = df['fecha'].apply(lambda x: x.split(' ')[0] if ' ' in x else x)
            filtro = df[df['fecha_simple'] == fecha_busqueda]
            ventas, total_dia = [], 0.0
            
            for index, row in filtro.iterrows():
                try: hora = row['fecha'].split(' ')[1][:5]
                except: hora = "00:00"
                
                nombres = str(row['items'])
                try:
                    nombres = ", ".join([p['nombre'] for p in ast.literal_eval(nombres)])
                except: pass

                ventas.append({'hora': hora, 'total': float(row['total']), 'items': nombres})
                total_dia += float(row['total'])
            return {'ventas': ventas, 'total': total_dia}
        except:
            return {'ventas': [], 'total': 0.0}

    def obtener_todo_inventario(self):
        try:
            return pd.read_excel(ARCHIVO_DB, sheet_name="Productos", dtype={'codigo': str}).to_dict('records')
        except:
            return []

    def agregar_producto(self, codigo, nombre, precio, stock):
        try:
            df = pd.read_excel(ARCHIVO_DB, sheet_name="Productos", dtype={'codigo': str})
            if str(codigo).strip() in df['codigo'].astype(str).str.strip().values:
                return False, "Código ya existe"
            nuevo = pd.DataFrame([{'codigo': str(codigo).strip(), 'nombre': nombre, 'precio': float(precio), 'stock': int(stock)}])
            df_final = pd.concat([df, nuevo], ignore_index=True)
            with pd.ExcelWriter(ARCHIVO_DB, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_final.to_excel(writer, sheet_name="Productos", index=False)
            return True, "Ok"
        except Exception as e:
            return False, str(e)

    def modificar_stock(self, codigo, nuevo_stock):
        try:
            wb = load_workbook(ARCHIVO_DB)
            ws = wb["Productos"]
            for fila in range(2, ws.max_row + 1):
                if str(ws.cell(row=fila, column=1).value).strip() == str(codigo).strip():
                    ws.cell(row=fila, column=4, value=int(nuevo_stock))
                    wb.save(ARCHIVO_DB)
                    return True
            return False
        except:
            return False

    def eliminar_producto(self, codigo):
        try:
            df = pd.read_excel(ARCHIVO_DB, sheet_name="Productos", dtype={'codigo': str})
            df_final = df[df['codigo'].astype(str).str.strip() != str(codigo).strip()]
            with pd.ExcelWriter(ARCHIVO_DB, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_final.to_excel(writer, sheet_name="Productos", index=False)
            return True
        except:
            return False